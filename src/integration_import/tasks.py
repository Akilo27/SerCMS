# tasks.py
import hashlib
import time
from celery import shared_task
from django.db.utils import OperationalError
from django.db.models import Sum
from shop.models import (
    Products,
    Categories,
    Manufacturers,
    Site,
    ProductsGallery,
    Valute,
)
from integration_import.models import (
    ImportCsv,
    UploadFromDiskImportCsv,
    YandexDiskIntegration,
)
import validators
import logging
import csv
import os
from django.core.exceptions import ObjectDoesNotExist
from shop.models import (
    StockAvailability,
    Storage,
)
from .models import ImportCsv, ImportCsvElement
from webmain.models import SettingsGlobale, Blogs, Seo
from django.contrib.sites.models import Site
import requests
from django.core.files.base import ContentFile, File
from PIL import Image
from io import BytesIO
import uuid
from django.core.files.uploadedfile import SimpleUploadedFile
import openpyxl
from django.core.files.storage import default_storage
from datetime import datetime
from django.utils.timezone import now


try:
    disk = YandexDiskIntegration.objects.filter(activate=True).first()
    YANDEX_DISK_API_TOKEN = f"{disk.key}"
except:
    YANDEX_DISK_API_TOKEN = (
        "y0_AgAAAABWBjSIAAxqKAAAAAEQgf8IAAC7Te4ZBaxFQ57PWDem-EE321A46w"
    )
logger = logging.getLogger(__name__)


def save_chunk(import_csv_instance, chunk, headers, import_type):
    # Сохранение части данных в XLSX формат
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Добавление заголовков
    sheet.append(headers)

    # Добавление строк из чанка
    for row in chunk:
        sheet.append(row)

    # Сохранение временного XLSX файла
    file_name = f"{uuid.uuid4()}.xlsx"
    chunk_file_path = f"/tmp/{file_name}"
    workbook.save(chunk_file_path)

    # Загрузка файла в Django FileField через SimpleUploadedFile
    with open(chunk_file_path, "rb") as f:
        chunk_file = SimpleUploadedFile(
            file_name,
            f.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # Подсчет количества строк (минус заголовок)
    quantity_rows = len(chunk)

    # Создание элемента импорта с файлом XLSX
    import_csv_element = ImportCsvElement.objects.create(
        import_csv=import_csv_instance, file=chunk_file, quantity=quantity_rows
    )
    import_csv_instance.elements.add(import_csv_element)
    return import_csv_element


def download_image(import_csv_id, image_url):
    import_xlsx = ImportCsv.objects.get(id=import_csv_id)

    # Проверка: загружается ли изображение с диска
    if not import_xlsx.uploadfromdisk or not import_xlsx.uploadfromdisk.imagelink:
        try:
            # Загружаем изображение по URL
            response = requests.get(image_url)
            response.raise_for_status()
        except requests.RequestException as e:
            print(f"Ошибка при загрузке изображения: {e}")
            return None, None  # Возвращаем None и для хеша

        # Проверка, что загруженный файл является изображением
        content_type = response.headers.get("Content-Type", "")
        if "image" not in content_type:
            print("Загруженный файл не является изображением.")
            return None, None  # Возвращаем None и для хеша

        try:
            # Открываем изображение
            image = Image.open(BytesIO(response.content))

            # Вычисляем хеш загруженного изображения
            image_hash = hashlib.md5(response.content).hexdigest()

            # Пытаемся конвертировать изображение в WebP
            webp_image = BytesIO()
            image.save(webp_image, format="WebP")
            webp_image.seek(0)

            # Генерация имени файла WebP
            original_filename = os.path.basename(image_url.split("?")[0])
            original_name_without_ext, _ = os.path.splitext(original_filename)
            webp_filename = f"{original_name_without_ext}.webp"

            print(
                f"Изображение успешно преобразовано в WebP и сохранено как: {webp_filename}"
            )
            return ContentFile(
                webp_image.read(), name=webp_filename
            ), image_hash  # Возвращаем изображение и хеш

        except Exception as e:
            print(f"Ошибка при преобразовании изображения в WebP: {e}")
            print("Сохранение изображения в исходном формате...")

            try:
                # Сохраняем исходное изображение
                original_format_image = BytesIO(response.content)
                original_filename = os.path.basename(image_url.split("?")[0])

                # Вычисляем хеш исходного изображения
                image_hash = hashlib.md5(response.content).hexdigest()

                print(
                    f"Изображение сохранено в исходном формате как: {original_filename}"
                )
                return ContentFile(
                    original_format_image.read(), name=original_filename
                ), image_hash  # Возвращаем исходное изображение и хеш
            except Exception as e:
                print(f"Ошибка при сохранении исходного изображения: {e}")
                return None, None  # Возвращаем None и для хеша

    else:
        # Загружаем изображение с Яндекс.Диска
        base_path = import_xlsx.uploadfromdisk.imagelink
        full_image_path = os.path.join(base_path, image_url)

        api_url = f"https://cloud-api.yandex.net/v1/disk/resources/download?path={full_image_path}"

        headers = {"Authorization": f"OAuth {YANDEX_DISK_API_TOKEN}"}

        try:
            # Получаем ссылку для скачивания файла с Яндекс.Диска
            response = requests.get(api_url, headers=headers)
            response.raise_for_status()

            download_url = response.json().get("href")

            if not download_url:
                print("Ошибка: не удалось получить ссылку для скачивания файла.")
                return None, None  # Возвращаем None и для хеша

            # Загружаем изображение по полученной ссылке
            image_response = requests.get(download_url)
            image_response.raise_for_status()

            # Проверяем, является ли файл изображением
            content_type = image_response.headers.get("Content-Type", "")
            if "image" not in content_type:
                print("Загруженный файл не является изображением.")
                return None, None  # Возвращаем None и для хеша

            try:
                # Открываем и конвертируем изображение в WebP
                image = Image.open(BytesIO(image_response.content))
                webp_image = BytesIO()
                image.save(webp_image, format="WebP")
                webp_image.seek(0)

                # Вычисляем хеш изображения
                image_hash = hashlib.md5(image_response.content).hexdigest()

                original_name_without_ext, _ = os.path.splitext(image_url)
                webp_filename = f"{original_name_without_ext}.webp"

                return ContentFile(
                    webp_image.read(), name=webp_filename
                ), image_hash  # Возвращаем изображение и хеш

            except Exception as e:
                print(f"Ошибка при преобразовании изображения в WebP: {e}")
                return None, None  # Возвращаем None и для хеша

        except requests.RequestException as e:
            print(f"Ошибка при загрузке изображения с Яндекс.Диска: {e}")
            return None, None  # Возвращаем None и для хеша


def to_boolean(value):
    if value == "1":
        return True
    elif value == "0":
        return False
    return False


@shared_task
def start_of_processing_csv(import_csv_id):
    instance = ImportCsv.objects.get(id=import_csv_id)
    if instance.status:
        return  # Задача уже завершена или все элементы созданы

    # Открываем исходный XLSX файл
    xlsx_file = instance.file
    xlsx_file.seek(0)
    workbook = openpyxl.load_workbook(xlsx_file, data_only=True)
    sheet = workbook.active

    # Получаем заголовки
    headers = [cell.value for cell in sheet[1]]
    chunk_size = 1000  # Размер части
    chunk = []

    row_limit = instance.quantity
    processed_rows = 0

    import_csv_elements = []

    # Обработка строк, начиная со второй (первая строка - заголовок)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if instance.created_all_elements:
            break
        chunk.append(row)
        processed_rows += 1

        if len(chunk) >= chunk_size or processed_rows >= row_limit:
            import_csv_element = save_chunk(instance, chunk, headers, instance.type)
            import_csv_elements.append(import_csv_element)
            chunk = []

        if processed_rows >= row_limit:
            break

    # Сохраняем оставшиеся данные
    if chunk:
        import_csv_element = save_chunk(instance, chunk, headers, instance.type)
        import_csv_elements.append(import_csv_element)

    # Устанавливаем флаг после создания всех ImportCsvElement
    instance.created_all_elements = True
    instance.save()

    # Запуск задач обработки для всех созданных элементов
    for element in import_csv_elements:
        if instance.type == 1:
            process_product_csv_task.delay(element.id, element.file.path)
        elif instance.type == 2:
            process_category_csv_task.delay(element.id, element.file.path)
        elif instance.type == 6:
            process_categories_csv_task.delay(element.id, element.file.path)
        elif instance.type == 4:
            process_manufacturer_csv_task.delay(element.id, element.file.path)
        elif instance.type == 5:
            process_blog_csv_task.delay(element.id, element.file.path)
        elif instance.type == 7:
            process_seo_csv_task.delay(element.id, element.file.path)
        elif instance.type == 9:
            process_banner_csv_task.delay(element.id, element.file.path)
        else:
            logger.error("Неизвестный тип импорта.", exc_info=True)


def safe_int(value):
    try:
        # Проверка, если значение пустое, то возвращаем 0
        return int(value) if value.strip() else 0
    except ValueError:
        return 0  # Если не удалось преобразовать, возвращаем 0


@shared_task
def process_product_csv_task(import_csv_id, xlsx_file_path):
    import_csv_instance = ImportCsvElement.objects.get(id=import_csv_id, status=False)
    import_csv = import_csv_instance.import_csv

    try:
        # Открываем файл XLSX
        with default_storage.open(xlsx_file_path, "rb") as file:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            header = [cell.value for cell in sheet[1]]  # Заголовки
            required_headers = [
                "id",
                "name",
                "price",
                "category",
                "manufacturers",
                "image",
                "order",
                "review",
                "faqs",
                "site",
                "valute",
            ]
            optional_headers = [
                "fragment",
                "description",
                "slug",
                "title",
                "content",
                "discount",
                "quantity",
                "review_rating",
                "type",
                "comment",
                "images",
                "atribute",
                "stock",
                "width",
                "height",
                "length",
                "weight",
            ]

            header_dict = {header[i]: i for i in range(len(header))}
            missing_headers = [
                field for field in required_headers if field not in header_dict
            ]

            if missing_headers:
                logger.error(
                    f'Отсутствуют обязательные заголовки: {", ".join(missing_headers)}'
                )
                return

            # Начинаем обработку данных
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    # Преобразуем значения в строки, чтобы избежать ошибок
                    product_data = {
                        field: str(row[header_dict[field]]).strip()
                        if field in header_dict and row[header_dict[field]] is not None
                        else None
                        for field in required_headers + optional_headers
                    }

                    if product_data["slug"] and "%" in product_data["slug"]:
                        product_data["slug"] = ""
                    try:
                        product_data["price"] = float(product_data["price"])
                    except:
                        logger.error("Товар не имеет цены следовательно будет пропущен")
                        continue

                    product_data["order"] = to_boolean(product_data["order"])
                    product_data["review"] = to_boolean(product_data["review"])
                    product_data["faqs"] = to_boolean(product_data["faqs"])
                    product_data["comment"] = to_boolean(
                        product_data.get("comment", "0")
                    )
                    product_data["width"] = safe_int(product_data.get("width", ""))
                    product_data["height"] = safe_int(product_data.get("height", ""))
                    product_data["length"] = safe_int(product_data.get("length", ""))
                    product_data["weight"] = safe_int(product_data.get("weight", ""))
                    sites = []
                    for site_id in product_data["site"].split("."):
                        try:
                            site = Site.objects.get(id=site_id.strip())
                            sites.append(site)
                        except ObjectDoesNotExist:
                            continue

                    settingsglobales = SettingsGlobale.objects.filter(site__in=sites)

                    for settingsglobale in settingsglobales:
                        if product_data["price"] > settingsglobale.max_price_product:
                            settingsglobale.max_price_product = product_data["price"]
                            settingsglobale.save()

                    categories = []
                    parent_category = None
                    for category_name in product_data["category"].split("|"):
                        category_name = category_name.strip()
                        try:
                            category = (
                                Categories.objects.filter(
                                    name=category_name,
                                    parent=parent_category,
                                    site__in=sites,
                                )
                                .distinct()
                                .get()
                            )
                        except Categories.DoesNotExist:
                            category = Categories.objects.create(
                                name=category_name, parent=parent_category
                            )
                            category.site.set(sites)
                        categories.append(category)
                        parent_category = category

                    manufacturer, created = Manufacturers.objects.get_or_create(
                        name=product_data["manufacturers"]
                    )
                    if created:
                        manufacturer.site.set(sites)

                    valute, created = Valute.objects.get_or_create(
                        key=product_data["valute"]
                    )

                    manufacturer_identifier = product_data["id"]

                    if not manufacturer_identifier:
                        logger.error(
                            f'Отсутствует идентификатор производителя для продукта с именем: {product_data["name"]} с id {product_data["id"]} полный дата {product_data}'
                        )
                        continue

                    type_mapping = {
                        "Вес": 2,
                        "Упк": 3,
                        "Обычный": 1,
                    }

                    typeint = 1  # Значение по умолчанию

                    if "type" in product_data:
                        type_value = product_data["type"]
                        logger.debug(f"Обрабатывается тип продукта: '{type_value}'")

                        if isinstance(type_value, str):
                            type_value = type_value.strip()
                            logger.debug(
                                f"Обработанное строковое значение типа продукта: '{type_value}'"
                            )
                            typeint = type_mapping.get(
                                type_value, 1
                            )  # Используем 1 как значение по умолчанию
                            logger.info(
                                f"Тип продукта '{type_value}' соответствует значению {typeint}"
                            )
                        elif isinstance(type_value, (int, float)):
                            typeint = int(type_value)
                            logger.info(
                                f"Использовано числовое значение типа продукта: {typeint}"
                            )
                        else:
                            logger.warning(
                                f"Неподдерживаемый тип данных для product_data['type']: {type(type_value)}. Используется значение по умолчанию (1)."
                            )
                    else:
                        logger.warning(
                            "Ключ 'type' отсутствует в product_data. Используется значение по умолчанию (1)."
                        )

                    logger.info(
                        f"Итоговое значение typeint перед сохранением: {typeint}"
                    )
                    product_get = Products.objects.filter(
                        manufacturer_identifier=manufacturer_identifier,
                        manufacturers=manufacturer,
                    ).first()
                    try:
                        if product_get:
                            if product_get.image:
                                # Получаем текущее имя файла изображения без расширения
                                current_image_filename = os.path.splitext(
                                    os.path.basename(product_get.image.name)
                                )[0]
                            else:
                                current_image_filename = None
                        else:
                            current_image_filename = None

                        # Скачиваем новое изображение
                        new_image_content, new_image_hash = download_image(
                            import_csv.id, product_data["image"]
                        )

                        # Извлекаем имя нового изображения без расширения
                        new_image_filename = (
                            os.path.splitext(os.path.basename(new_image_content.name))[
                                0
                            ]
                            if new_image_content
                            else None
                        )

                        # Сравниваем имена файлов без учета расширений
                        if new_image_filename != current_image_filename:
                            # Если имена различаются, обновляем изображение
                            product_data["image_content"] = new_image_content
                            print(
                                f"Обновлено изображение для продукта {product_data['id']} - новое имя файла: {new_image_filename} - cтарое {current_image_filename}"
                            )
                        else:
                            # Если имена одинаковы, используем текущее изображение
                            product_data["image_content"] = product_get.image
                            print(
                                f"Изображение для продукта {product_data['id']} не изменилось, загрузка не требуется."
                            )

                    except Exception as e:
                        if product_get and product_get.image:
                            product_data["image_content"] = (
                                product_get.image
                            )  # Используем текущее изображение в случае ошибки
                        else:
                            product_data["image_content"] = None
                        logger.info(f"Ошибка при обработке изображения: {e}")

                    product, product_created = Products.objects.update_or_create(
                        manufacturer_identifier=manufacturer_identifier,
                        manufacturers=manufacturer,
                        defaults={
                            "name": product_data["name"],
                            "typeofchoice": typeint,
                            "fragment": product_data["fragment"],
                            "description": product_data["description"],
                            "slug": product_data["slug"],
                            "title": product_data["title"],
                            "content": product_data["content"],
                            "image": product_data["image_content"],
                            "price": product_data["price"],
                            "discount": product_data["discount"],
                            "quantity": product_data["quantity"],
                            "review_rating": product_data["review_rating"],
                            "order": product_data["order"],
                            "review": product_data["review"],
                            "faqs": product_data["faqs"],
                            "comment": product_data["comment"],
                            "width": product_data["width"],
                            "height": product_data["height"],
                            "length": product_data["length"],
                            "weight": product_data["weight"],
                            "valute": valute,
                        },
                    )

                    if product_created:
                        import_csv_instance.added_element += 1
                        import_csv_instance.save()
                    else:
                        import_csv_instance.upload_element += 1
                        import_csv_instance.save()

                    product.category.set(categories)
                    product.site.set(sites)

                    if product_data["images"]:
                        for image_url in product_data["images"].split(","):
                            try:
                                # Загрузка изображения и вычисление хеша
                                image_content, image_hash = download_image(
                                    import_csv.id, image_url.strip()
                                )

                                if image_content and image_hash:
                                    # Проверяем, существует ли уже такое изображение для данного продукта по хешу
                                    existing_image = ProductsGallery.objects.filter(
                                        products=product, image_hash=image_hash
                                    ).first()

                                    if not existing_image:
                                        # Если изображение с таким хешем не найдено, создаем новую запись в ProductsGallery
                                        gallery = ProductsGallery(
                                            products=product, image_hash=image_hash
                                        )
                                        gallery.image.save(
                                            image_content.name, image_content, save=True
                                        )
                                        print(
                                            f"Добавлено новое изображение для продукта {product.id}: {image_content.name}"
                                        )
                                    else:
                                        print(
                                            f"Изображение с хешем {image_hash} уже существует для продукта {product.id}, пропуск загрузки."
                                        )

                            except Exception as e:
                                print(
                                    f"Ошибка при обработке изображения {image_url.strip()}: {e}"
                                )
                                continue

                    if product_data["atribute"]:
                        for pair in product_data["atribute"].split("|"):
                            if not pair.strip():
                                continue
                            try:
                                storage_id, quantity = map(int, pair.split(":"))
                                storage = Storage.objects.get(id=storage_id)
                                StockAvailability.objects.get_or_create(
                                    products=product, storage=storage, quantity=quantity
                                )
                            except (ValueError, ObjectDoesNotExist):
                                continue

                    if product_data["stock"]:
                        for pair in product_data["stock"].split(","):
                            if not pair.strip():
                                continue
                            try:
                                storage_id, quantity = map(int, pair.split(":"))
                                storage = Storage.objects.get(id=storage_id)
                                StockAvailability.objects.get_or_create(
                                    products=product, storage=storage, quantity=quantity
                                )
                            except (ValueError, ObjectDoesNotExist):
                                continue

                    import_csv_instance.passed += 1
                    import_csv_instance.save()

                    import_csv.passed = import_csv.elements.aggregate(Sum("passed"))[
                        "passed__sum"
                    ]
                    import_csv.added_element = import_csv.elements.aggregate(
                        Sum("added_element")
                    )["added_element__sum"]
                    import_csv.upload_element = import_csv.elements.aggregate(
                        Sum("upload_element")
                    )["upload_element__sum"]
                    import_csv.save()

                except OperationalError as e:
                    logger.error(f"Database is locked: {e}. Retrying...", exc_info=True)
                    time.sleep(1)
                except (IndexError, ValueError) as e:
                    logger.error(f"Ошибка при обработке строки: {e}", exc_info=True)
                except Exception as e:
                    logger.error(f"Произошла ошибка: {e}", exc_info=True)

        import_csv_instance.status = True
        import_csv_instance.save()
        import_csv.save()

    except Exception as e:
        import_csv_instance.status = False
        import_csv_instance.save()
        logger.error(f"Произошла ошибка при обработке XLSX файла: {e}", exc_info=True)


@shared_task
def process_category_csv_task(import_csv_id, csv_file_content):
    import_csv_element = ImportCsvElement.objects.get(id=import_csv_id)
    import_csv = import_csv_element.import_csv

    def process_csv():
        try:
            with open(csv_file_content, "r", encoding="UTF-8") as file:
                data_set = file.read()
            csv_reader = csv.reader(data_set.splitlines(), delimiter=";")

            try:
                header = next(csv_reader)
            except StopIteration:
                logger.error("Пустой CSV файл.")
                import_csv_element.status = False
                import_csv_element.save()
                return

            required_headers = [
                "id",
                "name",
                "description",
                "title",
                "content",
                "slug",
                "site",
                "cover",
                "icon",
                "image",
                "parent",
            ]
            missing_headers = [h for h in required_headers if h not in header]
            if missing_headers:
                logger.error(f"Отсутствуют необходимые заголовки: {missing_headers}")
                import_csv_element.status = False
                import_csv_element.save()
                return

            passed_count = 0
            categories = []

            for row in csv_reader:
                try:
                    category_data = {
                        field: row[header.index(field)].strip()
                        if field in header
                        else None
                        for field in required_headers
                    }

                    # Получение сайтов
                    site_ids = category_data["site"].split(",")
                    sites = []
                    for site_id in site_ids:
                        try:
                            site = Site.objects.get(id=site_id.strip())
                            sites.append(site)
                        except Site.DoesNotExist:
                            logger.warning(f"Сайт с ID {site_id} не найден. Пропущено.")

                    # Получение родительской категории
                    parent_category = None
                    parent_id = category_data.get("parent")
                    if parent_id:
                        try:
                            parent_category = Categories.objects.get(
                                id=parent_id.strip()
                            )
                        except Categories.DoesNotExist:
                            logger.warning(
                                f"Родительская категория с ID {parent_id} не найдена. Пропущено."
                            )

                    # Загрузка изображений
                    for field in ["cover", "icon", "image"]:
                        if field not in category_data or not category_data[field]:
                            continue
                        image_url = category_data[field]
                        if not validators.url(image_url):
                            logger.warning(
                                f"Неверный URL изображения ({field}): {image_url}"
                            )
                            continue
                        try:
                            image_content = download_image(import_csv.id, image_url)
                            if image_content is None:
                                logger.warning(
                                    f"Не удалось загрузить изображение ({field}): {image_url}"
                                )
                                continue
                        except Exception as e:
                            logger.warning(
                                f"Ошибка при загрузке изображения ({field}): {e}"
                            )
                            continue
                        category_data[f"{field}_content"] = image_content

                    # Создание или обновление категории
                    category_defaults = {
                        "name": category_data["name"],
                        "description": category_data["description"],
                        "title": category_data["title"],
                        "content": category_data["content"],
                        "slug": category_data["slug"],
                        "parent": parent_category,
                    }
                    category, created = Categories.objects.update_or_create(
                        id=category_data["id"], defaults=category_defaults
                    )

                    if created:
                        import_csv_element.added_element += 1
                        import_csv_element.save()
                    else:
                        import_csv_element.upload_element += 1
                        import_csv_element.save()

                    # Сохраняем информацию о категории (category, created) в списке categories
                    categories.append((category, created))

                    # Сохранение изображений
                    for field in ["cover", "icon", "image"]:
                        if f"{field}_content" in category_data:
                            getattr(category, field).save(
                                f"{category_data['id']}_{field}.jpg",
                                category_data[f"{field}_content"],
                                save=True,
                            )

                    # Установка сайтов
                    category.site.set(sites)

                    import_csv_element.passed += 1
                    import_csv_element.save()

                    import_csv.passed = import_csv.elements.aggregate(Sum("passed"))[
                        "passed__sum"
                    ]
                    import_csv.added_element = import_csv.elements.aggregate(
                        Sum("added_element")
                    )["added_element__sum"]
                    import_csv.upload_element = import_csv.elements.aggregate(
                        Sum("upload_element")
                    )["upload_element__sum"]
                    import_csv.save()

                    if created:
                        logger.info(f"Создана новая категория: {category_data['id']}")
                    else:
                        logger.info(
                            f"Обновлена существующая категория: {category_data['id']}"
                        )
                except (IndexError, ValueError) as e:
                    logger.error(f"Ошибка при обработке строки: {e}")
                except Exception as e:
                    logger.error(f"Произошла ошибка: {e}", exc_info=True)

            import_csv_element.status = True
            import_csv_element.save()
            all_elements_have_status = (
                import_csv.elements.all().aggregate(all_status_true=Sum("status"))[
                    "all_status_true"
                ]
                == import_csv.elements.count()
            )
            if all_elements_have_status:
                import_csv.status = True
                import_csv.elements.all().delete()
                import_csv.save()
        except Exception as e:
            import_csv_element.status = False
            import_csv_element.save()
            logger.error(f"Произошла ошибка при обработке CSV: {e}", exc_info=True)

    process_csv()


@shared_task
def process_categories_csv_task(import_csv_id, csv_file_content):
    import_csv_element = ImportCsvElement.objects.get(id=import_csv_id)
    import_csv = import_csv_element.import_csv

    def process_csv():
        try:
            with open(csv_file_content, "r", encoding="UTF-8") as file:
                data_set = file.read()
            csv_reader = csv.reader(data_set.splitlines(), delimiter=";")

            try:
                header = next(csv_reader)
            except StopIteration:
                logger.error("Пустой CSV файл.")
                import_csv_element.status = False
                import_csv_element.save()
                return

            required_headers = [
                "id",
                "name",
                "description",
                "title",
                "content",
                "slug",
                "site",
                "cover",
                "icon",
                "image",
                "parent",
            ]
            missing_headers = [h for h in required_headers if h not in header]
            if missing_headers:
                logger.error(f"Отсутствуют необходимые заголовки: {missing_headers}")
                import_csv_element.status = False
                import_csv_element.save()
                return

            passed_count = 0
            categories = []

            for row in csv_reader:
                try:
                    category_data = {
                        field: row[header.index(field)].strip()
                        if field in header
                        else None
                        for field in required_headers
                    }

                    # Получение сайтов
                    site_ids = category_data["site"].split(",")
                    sites = []
                    for site_id in site_ids:
                        try:
                            site = Site.objects.get(id=site_id.strip())
                            sites.append(site)
                        except Site.DoesNotExist:
                            logger.warning(f"Сайт с ID {site_id} не найден. Пропущено.")

                    # Получение родительской категории
                    parent_category = None
                    parent_id = category_data.get("parent")
                    if parent_id:
                        try:
                            parent_category = Categories.objects.get(
                                id=parent_id.strip()
                            )
                        except Categories.DoesNotExist:
                            logger.warning(
                                f"Родительская категория с ID {parent_id} не найдена. Пропущено."
                            )

                    # Загрузка изображений
                    for field in ["cover", "icon", "image"]:
                        if field not in category_data or not category_data[field]:
                            continue
                        image_url = category_data[field]
                        if not validators.url(image_url):
                            logger.warning(
                                f"Неверный URL изображения ({field}): {image_url}"
                            )
                            continue  # Skip processing this row
                        try:
                            image_content = download_image(import_csv.id, image_url)
                            if image_content is None:
                                logger.warning(
                                    f"Не удалось загрузить изображение ({field}): {image_url}"
                                )
                                continue
                        except Exception as e:
                            logger.warning(
                                f"Ошибка при загрузке изображения ({field}): {e}"
                            )
                            continue
                        category_data[f"{field}_content"] = image_content

                    # Создание или обновление категории
                    category_defaults = {
                        "name": category_data["name"],
                        "description": category_data["description"],
                        "title": category_data["title"],
                        "content": category_data["content"],
                        "slug": category_data["slug"],
                        "parent": parent_category,
                    }
                    category, created = Categories.objects.update_or_create(
                        id=category_data["id"], defaults=category_defaults
                    )

                    if created:
                        import_csv_element.added_element += 1
                        import_csv_element.save()
                    else:
                        import_csv_element.upload_element += 1
                        import_csv_element.save()

                    # Сохраняем информацию о категории (category, created) в списке categories
                    categories.append((category, created))

                    # Сохранение изображений
                    for field in ["cover", "icon", "image"]:
                        if f"{field}_content" in category_data:
                            getattr(category, field).save(
                                f"{category_data['id']}_{field}.jpg",
                                category_data[f"{field}_content"],
                                save=True,
                            )

                    # Установка сайтов
                    category.site.set(sites)

                    import_csv_element.passed += 1
                    import_csv_element.save()

                    import_csv.passed = import_csv.elements.aggregate(Sum("passed"))[
                        "passed__sum"
                    ]
                    import_csv.added_element = import_csv.elements.aggregate(
                        Sum("added_element")
                    )["added_element__sum"]
                    import_csv.upload_element = import_csv.elements.aggregate(
                        Sum("upload_element")
                    )["upload_element__sum"]
                    import_csv.save()
                    if created:
                        logger.info(f"Создана новая категория: {category_data['id']}")
                    else:
                        logger.info(
                            f"Обновлена существующая категория: {category_data['id']}"
                        )
                except (IndexError, ValueError) as e:
                    logger.error(f"Ошибка при обработке строки: {e}")
                except Exception as e:
                    logger.error(f"Произошла ошибка: {e}", exc_info=True)

            # Обновить поле added_element и upload_element в экземпляре import_csv после обработки

            import_csv_element.status = True
            import_csv_element.save()

            all_elements_have_status = (
                import_csv.elements.all().aggregate(all_status_true=Sum("status"))[
                    "all_status_true"
                ]
                == import_csv.elements.count()
            )
            if all_elements_have_status:
                import_csv.status = True
                import_csv.elements.all().delete()
                import_csv.save()

        except Exception as e:
            import_csv_element.status = False
            import_csv_element.save()
            logger.error(f"Произошла ошибка при обработке CSV: {e}", exc_info=True)

    process_csv()


@shared_task
def process_manufacturer_csv_task(import_csv_id, csv_file_content):
    import_csv_element = ImportCsvElement.objects.get(id=import_csv_id)
    import_csv = import_csv_element.import_csv

    def process_csv():
        try:
            with open(csv_file_content, "r", encoding="UTF-8") as file:
                data_set = file.read()
            csv_reader = csv.reader(data_set.splitlines(), delimiter=";")

            try:
                header = next(csv_reader)
            except StopIteration:
                logger.error("Пустой CSV файл.")
                import_csv_element.status = False
                import_csv_element.save()
                return

            required_headers = [
                "id",
                "name",
                "description",
                "title",
                "content",
                "slug",
                "site",
                "cover",
                "image",
            ]
            missing_headers = [h for h in required_headers if h not in header]
            if missing_headers:
                logger.error(f"Отсутствуют необходимые заголовки: {missing_headers}")
                import_csv_element.status = False
                import_csv_element.save()
                return

            passed_count = 0
            manufacturers = []

            for row in csv_reader:
                try:
                    manufacturer_data = {
                        field: row[header.index(field)].strip()
                        if field in header
                        else None
                        for field in required_headers
                    }

                    # Получение сайтов
                    site_ids = manufacturer_data["site"].split(",")
                    sites = []
                    for site_id in site_ids:
                        try:
                            site = Site.objects.get(id=site_id.strip())
                            sites.append(site)
                        except ObjectDoesNotExist:
                            logger.warning(f"Сайт с ID {site_id} не найден. Пропущено.")

                    # Загрузка изображений
                    for field in ["cover", "image"]:
                        if (
                            field not in manufacturer_data
                            or not manufacturer_data[field]
                        ):
                            continue
                        image_url = manufacturer_data[field]
                        if not validators.url(image_url):
                            logger.warning(
                                f"Неверный URL изображения ({field}): {image_url}"
                            )
                            continue
                        try:
                            image_content = download_image(import_csv.id, image_url)
                            if image_content is None:
                                logger.warning(
                                    f"Не удалось загрузить изображение ({field}): {image_url}"
                                )
                                continue
                        except Exception as e:
                            logger.warning(
                                f"Ошибка при загрузке изображения ({field}): {e}"
                            )
                            continue
                        manufacturer_data[f"{field}_content"] = image_content

                    # Создание или обновление производителя
                    manufacturer_defaults = {
                        "name": manufacturer_data["name"],
                        "description": manufacturer_data["description"],
                        "title": manufacturer_data["title"],
                        "content": manufacturer_data["content"],
                        "slug": manufacturer_data["slug"]
                        if manufacturer_data["slug"]
                        else None,
                    }
                    manufacturer, created = Manufacturers.objects.update_or_create(
                        id=manufacturer_data["id"], defaults=manufacturer_defaults
                    )

                    if created:
                        import_csv_element.added_element += 1
                        import_csv_element.save()
                    else:
                        import_csv_element.upload_element += 1
                        import_csv_element.save()

                    # Сохранение информации о производителе (manufacturer, created) в списке manufacturers
                    manufacturers.append((manufacturer, created))

                    # Сохранение изображений
                    for field in ["cover", "image"]:
                        if f"{field}_content" in manufacturer_data:
                            getattr(manufacturer, field).save(
                                f"{manufacturer_data['id']}_{field}.jpg",
                                manufacturer_data[f"{field}_content"],
                                save=True,
                            )

                    # Установка сайтов
                    manufacturer.site.set(sites)

                    import_csv_element.passed += 1
                    import_csv_element.save()

                    import_csv.passed = import_csv.elements.aggregate(Sum("passed"))[
                        "passed__sum"
                    ]
                    import_csv.added_element = import_csv.elements.aggregate(
                        Sum("added_element")
                    )["added_element__sum"]
                    import_csv.upload_element = import_csv.elements.aggregate(
                        Sum("upload_element")
                    )["upload_element__sum"]
                    import_csv.save()
                    if created:
                        logger.info(
                            f"Создан новый производитель: {manufacturer_data['id']}"
                        )
                    else:
                        logger.info(
                            f"Обновлен существующий производитель: {manufacturer_data['id']}"
                        )
                except (IndexError, ValueError) as e:
                    logger.error(f"Ошибка при обработке строки: {e}")
                except Exception as e:
                    logger.error(f"Произошла ошибка: {e}", exc_info=True)

            # Обновление поля added_element и upload_element в экземпляре import_csv после обработки
            import_csv_element.status = True
            import_csv_element.save()

            all_elements_have_status = (
                import_csv.elements.all().aggregate(all_status_true=Sum("status"))[
                    "all_status_true"
                ]
                == import_csv.elements.count()
            )
            if all_elements_have_status:
                import_csv.status = True
                import_csv.elements.all().delete()
                import_csv.save()

        except Exception as e:
            import_csv_element.status = False
            import_csv_element.save()
            logger.error(f"Произошла ошибка при обработке CSV: {e}", exc_info=True)

    process_csv()


@shared_task
def process_blog_csv_task(import_csv_id, csv_file_content):
    import_csv_element = ImportCsvElement.objects.get(id=import_csv_id)
    import_csv = import_csv_element.import_csv

    def process_csv():
        try:
            with open(csv_file_content, "r", encoding="UTF-8") as file:
                data_set = file.read()
            csv_reader = csv.reader(data_set.splitlines(), delimiter=";")

            try:
                header = next(csv_reader)
            except StopIteration:
                logger.error("Пустой CSV файл.")
                import_csv_element.status = False
                import_csv_element.save()
                return

            required_headers = [
                "id",
                "site",
                "author",
                "resource",
                "category",
                "name",
                "description",
                "previev",
                "title",
                "content",
                "propertytitle",
                "propertydescription",
                "slug",
                "cover",
            ]
            missing_headers = [h for h in required_headers if h not in header]
            if missing_headers:
                logger.error(f"Отсутствуют необходимые заголовки: {missing_headers}")
                import_csv_element.status = False
                import_csv_element.save()
                return

            optional_headers = []

            passed_count = 0
            blogss = []

            for row in csv_reader:
                try:
                    blogs_data = {}
                    for field in required_headers + optional_headers:
                        if field in header:
                            blogs_data[field] = row[header.index(field)].strip()
                        else:
                            blogs_data[field] = None

                    if blogs_data["slug"] and "%" in blogs_data["slug"]:
                        logger.info("Значение 'slug' содержит '%', оставляем пустым.")
                        blogs_data["slug"] = ""

                    try:
                        blogs_data["image_content"] = download_image(
                            import_csv.id, blogs_data["cover"]
                        )
                        if blogs_data["image_content"] is None:
                            logger.warning(
                                f"Не удалось загрузить изображение: {blogs_data['cover']}"
                            )
                    except Exception as e:
                        logger.warning(f"Ошибка при загрузке изображения: {e}")
                        blogs_data["image_content"] = None

                    try:
                        blogs_data["previev_content"] = download_image(
                            import_csv.id, blogs_data["previev"]
                        )
                        if blogs_data["previev_content"] is None:
                            logger.warning(
                                f"Не удалось загрузить изображение: {blogs_data['previev']}"
                            )
                    except Exception as e:
                        logger.warning(f"Ошибка при загрузке изображения: {e}")
                        blogs_data["previev_content"] = None

                    site_ids = blogs_data["site"].split(",")
                    sites = []
                    for site_id in site_ids:
                        try:
                            site = Site.objects.get(id=site_id.strip())
                            sites.append(site)
                        except ObjectDoesNotExist:
                            logger.warning(f"Сайт с ID {site_id} не найден. Пропущено.")

                    category_names = blogs_data["category"].split("|")
                    categories = []
                    parent_category = None
                    for category_name in category_names:
                        category_name = category_name.strip()
                        if not parent_category:
                            parent_category, created = Categorys.objects.get_or_create(
                                name=category_name
                            )
                            if created:
                                parent_category.site.set(sites)
                        else:
                            child_category, created = Categorys.objects.get_or_create(
                                name=category_name, defaults={"parent": parent_category}
                            )
                            if created:
                                child_category.site.set(sites)
                            parent_category = child_category
                        categories.append(parent_category)

                    blogs, created = Blogs.objects.get_or_create(
                        id=blogs_data["id"],
                        defaults={
                            "author_id": blogs_data["author"]
                            if blogs_data["author"]
                            else None,
                            "resource": blogs_data["resource"],
                            "name": blogs_data["name"],
                            "description": blogs_data["description"],
                            "title": blogs_data["title"],
                            "content": blogs_data["content"],
                            "cover": blogs_data["image_content"],
                            "previev": blogs_data["previev_content"],
                            "propertytitle": blogs_data["propertytitle"],
                            "propertydescription": blogs_data["propertydescription"],
                            "slug": blogs_data["slug"],
                        },
                    )
                    if created:
                        import_csv_element.added_element += 1
                        import_csv_element.save()
                    else:
                        import_csv_element.upload_element += 1
                        import_csv_element.save()
                    blogss.append((blogs, created))

                    if created:
                        logger.info(f"Создан новый продукт: {blogs_data['id']}")
                    else:
                        logger.info(
                            f"Обновлен существующий продукт: {blogs_data['id']}"
                        )
                        blogs.author = blogs_data["author"]
                        blogs.resource = blogs_data["resource"]
                        blogs.name = blogs_data["name"]
                        blogs.description = blogs_data["description"]
                        blogs.title = blogs_data["title"]
                        blogs.content = blogs_data["content"]
                        blogs.cover = blogs_data["image_content"]
                        blogs.previev = blogs_data["previev_content"]
                        blogs.propertytitle = blogs_data["propertytitle"]
                        blogs.propertydescription = blogs_data["propertydescription"]
                        blogs.slug = blogs_data["slug"]
                        blogs.save()

                    blogs.category.set(categories)
                    blogs.site.set(sites)

                    import_csv_element.passed += 1
                    import_csv_element.save()

                    import_csv.passed = import_csv.elements.aggregate(Sum("passed"))[
                        "passed__sum"
                    ]
                    import_csv.added_element = import_csv.elements.aggregate(
                        Sum("added_element")
                    )["added_element__sum"]
                    import_csv.upload_element = import_csv.elements.aggregate(
                        Sum("upload_element")
                    )["upload_element__sum"]
                    import_csv.save()

                except (IndexError, ValueError) as e:
                    logger.error(f"Ошибка при обработке строки: {e}")
                except Exception as e:
                    logger.error(f"Произошла ошибка: {e}", exc_info=True)

            import_csv_element.status = True
            import_csv_element.save()

            all_elements_have_status = (
                import_csv.elements.all().aggregate(all_status_true=Sum("status"))[
                    "all_status_true"
                ]
                == import_csv.elements.count()
            )
            if all_elements_have_status:
                import_csv.status = True
                import_csv.elements.all().delete()
                import_csv.save()

        except Exception as e:
            import_csv_element.status = False
            import_csv_element.save()
            logger.error(f"Произошла ошибка при обработке CSV: {e}", exc_info=True)

    process_csv()


@shared_task
def process_seo_csv_task(import_csv_id, csv_file_content):
    import_csv_element = ImportCsvElement.objects.get(id=import_csv_id)
    import_csv = import_csv_element.import_csv

    def process_csv():
        try:
            with open(csv_file_content, "r", encoding="UTF-8") as file:
                data_set = file.read()
            csv_reader = csv.reader(data_set.splitlines(), delimiter=";")

            try:
                header = next(csv_reader)
            except StopIteration:
                logger.error("Пустой CSV файл.")
                import_csv_element.status = False
                import_csv_element.save()
                return

            required_headers = [
                "id",
                "setting",
                "pagetype",
                "previev",
                "title",
                "description",
                "propertytitle",
                "propertydescription",
            ]
            missing_headers = [h for h in required_headers if h not in header]
            if missing_headers:
                logger.error(f"Отсутствуют необходимые заголовки: {missing_headers}")
                import_csv_element.status = False
                import_csv_element.save()
                return

            optional_headers = []

            passed_count = 0
            seos = []

            for row in csv_reader:
                try:
                    seo_data = {}
                    for field in required_headers + optional_headers:
                        if field in header:
                            seo_data[field] = row[header.index(field)].strip()
                        else:
                            seo_data[field] = None

                    try:
                        seo_data["previev_content"] = download_image(
                            import_csv.id, seo_data["previev"]
                        )
                        if seo_data["previev_content"] is None:
                            logger.warning(
                                f"Не удалось загрузить изображение: {seo_data['previev']}"
                            )
                    except Exception as e:
                        logger.warning(f"Ошибка при загрузке изображения: {e}")
                        seo_data["previev_content"] = None

                    setting_id = int(seo_data["setting"])
                    settings_globale = SettingsGlobale.objects.get(id=setting_id)

                    seo, created = Seo.objects.get_or_create(
                        id=seo_data["id"],
                        defaults={
                            "setting": settings_globale,
                            "pagetype": seo_data["pagetype"],
                            "previev": seo_data["previev_content"],
                            "title": seo_data["title"],
                            "description": seo_data["description"],
                            "propertytitle": seo_data["propertytitle"],
                            "propertydescription": seo_data["propertydescription"],
                        },
                    )
                    if created:
                        import_csv_element.added_element += 1
                        import_csv_element.save()
                    else:
                        import_csv_element.upload_element += 1
                        import_csv_element.save()
                    seos.append((seo, created))

                    if created:
                        logger.info(f"Создан новый продукт: {seo_data['id']}")
                    else:
                        logger.info(f"Обновлен существующий продукт: {seo_data['id']}")
                        seo.pagetype = seo_data["pagetype"]
                        seo.setting = settings_globale
                        seo.previev = seo_data["previev_content"]
                        seo.title = seo_data["title"]
                        seo.description = seo_data["description"]
                        seo.propertytitle = seo_data["propertytitle"]
                        seo.propertydescription = seo_data["propertydescription"]
                        seo.save()

                    import_csv_element.passed += 1
                    import_csv_element.save()

                    import_csv.passed = import_csv.elements.aggregate(Sum("passed"))[
                        "passed__sum"
                    ]
                    import_csv.added_element = import_csv.elements.aggregate(
                        Sum("added_element")
                    )["added_element__sum"]
                    import_csv.upload_element = import_csv.elements.aggregate(
                        Sum("upload_element")
                    )["upload_element__sum"]
                    import_csv.save()

                except (IndexError, ValueError) as e:
                    logger.error(f"Ошибка при обработке строки: {e}")
                except Exception as e:
                    logger.error(f"Произошла ошибка: {e}", exc_info=True)
            import_csv_element.status = True
            import_csv_element.save()

            all_elements_have_status = (
                import_csv.elements.all().aggregate(all_status_true=Sum("status"))[
                    "all_status_true"
                ]
                == import_csv.elements.count()
            )
            if all_elements_have_status:
                import_csv.status = True
                import_csv.elements.all().delete()
                import_csv.save()

        except Exception as e:
            import_csv_element.status = False
            import_csv_element.save()
            logger.error(f"Произошла ошибка при обработке CSV: {e}", exc_info=True)

    process_csv()


@shared_task
def process_banner_csv_task(import_csv_id, csv_file_content):
    import_csv_element = ImportCsvElement.objects.get(id=import_csv_id)
    import_csv = import_csv_element.import_csv

    def process_csv():
        try:
            with open(csv_file_content, "r", encoding="UTF-8") as file:
                data_set = file.read()
            csv_reader = csv.reader(data_set.splitlines(), delimiter=";")

            try:
                header = next(csv_reader)
            except StopIteration:
                logger.error("Пустой CSV файл.")
                import_csv_element.status = False
                import_csv_element.save()
                return

            required_headers = ["id", "type", "image", "slug", "site"]
            missing_headers = [h for h in required_headers if h not in header]
            if missing_headers:
                logger.error(f"Отсутствуют необходимые заголовки: {missing_headers}")
                import_csv_element.status = False
                import_csv_element.save()
                return

            optional_headers = []

            passed_count = 0
            banners = []

            for row in csv_reader:
                try:
                    banner_data = {}
                    for field in required_headers + optional_headers:
                        if field in header:
                            banner_data[field] = row[header.index(field)].strip()
                        else:
                            banner_data[field] = None

                    try:
                        banner_data["image"] = download_image(
                            import_csv.id, banner_data["image"]
                        )
                        if banner_data["image"] is None:
                            logger.warning(
                                f"Не удалось загрузить изображение: {banner_data['image']}"
                            )
                    except Exception as e:
                        logger.warning(f"Ошибка при загрузке изображения: {e}")
                        banner_data["image"] = None

                    site_ids = banner_data["site"].split(",")
                    sites = []
                    for site_id in site_ids:
                        try:
                            site = Site.objects.get(id=site_id.strip())
                            sites.append(site)
                        except ObjectDoesNotExist:
                            logger.warning(f"Сайт с ID {site_id} не найден. Пропущено.")

                    banner, created = Banner.objects.get_or_create(
                        id=banner_data["id"],
                        defaults={
                            "type": banner_data["type"],
                            "image": banner_data["image"],
                            "slug": banner_data["slug"],
                        },
                    )
                    if created:
                        import_csv_element.added_element += 1
                        import_csv_element.save()
                    else:
                        import_csv_element.upload_element += 1
                        import_csv_element.save()
                    banners.append((banner, created))

                    if created:
                        logger.info(f"Создан новый баннер: {banner_data['id']}")
                    else:
                        logger.info(
                            f"Обновлен существующий баннер: {banner_data['id']}"
                        )
                        banner.type = banner_data["type"]
                        banner.image = banner_data["image"]
                        banner.slug = banner_data["slug"]
                        banner.save()

                    banner.site.set(sites)

                    import_csv_element.passed += 1
                    import_csv_element.save()

                    import_csv.passed = import_csv.elements.aggregate(Sum("passed"))[
                        "passed__sum"
                    ]
                    import_csv.added_element = import_csv.elements.aggregate(
                        Sum("added_element")
                    )["added_element__sum"]
                    import_csv.upload_element = import_csv.elements.aggregate(
                        Sum("upload_element")
                    )["upload_element__sum"]
                    import_csv.save()

                except (IndexError, ValueError) as e:
                    logger.error(f"Ошибка при обработке строки: {e}")
                except Exception as e:
                    logger.error(f"Произошла ошибка: {e}", exc_info=True)

            import_csv_element.status = True
            import_csv_element.save()

            all_elements_have_status = (
                import_csv.elements.all().aggregate(all_status_true=Sum("status"))[
                    "all_status_true"
                ]
                == import_csv.elements.count()
            )
            if all_elements_have_status:
                import_csv.status = True
                import_csv.elements.all().delete()
                import_csv.save()

        except Exception as e:
            import_csv_element.status = False
            import_csv_element.save()
            logger.error(f"Произошла ошибка при обработке CSV: {e}", exc_info=True)

    process_csv()


@shared_task
def download_file_from_model(upload_id, type):
    # type = 1: автоматическая загрузка
    # type = 2: ручная загрузка

    # Получаем объект с блокировкой транзакции
    upload_obj = UploadFromDiskImportCsv.objects.filter(
        id=upload_id, activate=True
    ).first()

    if not upload_obj:
        return f"Ошибка: объект с ID {upload_id} не найден или неактивен."

    # Проверка для автоматического типа (type = 1)
    if type == 1:
        if upload_obj.last_attempt and upload_obj.last_attempt.date() == now().date():
            return f"Загрузка уже выполнена сегодня для объекта с ID {upload_id}.upload_obj.last_attemp {upload_obj.last_attempt};upload_obj.last_attempt.date() {upload_obj.last_attempt.date() };now().date() {now().date()}"
        else:
            # Обновляем дату последней попытки
            upload_obj.last_attempt = now()
            upload_obj.save()

    # Продолжаем загрузку для любого типа
    file_path = upload_obj.link

    # Получаем URL для скачивания через API Яндекс.Диска
    api_url = (
        f"https://cloud-api.yandex.net/v1/disk/resources/download?path={file_path}"
    )
    headers = {"Authorization": f"OAuth {YANDEX_DISK_API_TOKEN}"}

    response = requests.get(api_url, headers=headers)
    response.raise_for_status()

    download_url = response.json().get("href")
    if not download_url:
        return "Ошибка: не удалось получить ссылку на скачивание."

    # Скачиваем файл по полученной ссылке
    download_response = requests.get(download_url, stream=True)
    download_response.raise_for_status()

    # Формируем путь для сохранения файла с текущей датой
    current_date = datetime.now().strftime("%Y/%m/%d")
    save_dir = os.path.join("import", current_date)
    os.makedirs(save_dir, exist_ok=True)

    save_path = os.path.join(save_dir, f"{upload_obj.type}_import_file.xlsx")

    # Сохраняем файл на диске
    with open(save_path, "wb") as file:
        for chunk in download_response.iter_content(chunk_size=8192):
            file.write(chunk)

    try:
        # Чтение Excel файла
        workbook = openpyxl.load_workbook(save_path, data_only=True)
        sheet = workbook.active
        row_count = sheet.max_row - 1  # Вычитание 1 для исключения заголовка
    except Exception as e:
        return f"Ошибка при чтении файла XLSX: {e}"

    # Открываем сохраненный файл для загрузки в базу данных
    with open(save_path, "rb") as f:
        file_content = File(f)  # Преобразуем в объект File для корректного сохранения

        # Создаем запись в базе данных с файлом
        ImportCsv.objects.create(
            type=upload_obj.type,
            status=False,
            quantity=row_count,
            added_element=0,
            upload_element=0,
            passed=0,
            file=file_content,  # Поле должно быть FileField в модели
            uploadfromdisk=upload_obj,
        )

    return f"Файл успешно скачан для объекта с ID {upload_id}."


@shared_task
def delete_import_from_db(importcsv_id):
    importcsv_obj = ImportCsv.objects.get(id=importcsv_id)
    if not importcsv_obj.file:
        return f"No file associated with ImportCsv id {importcsv_id}"
    file_path = importcsv_obj.file.path
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        manufacturer_ids = []
        header = [cell.value for cell in sheet[1]]
        id_index = header.index("id")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            manufacturer_ids.append(row[id_index])
    except Exception as e:
        return f"Error reading xlsx file: {str(e)}"
    if manufacturer_ids:
        deleted, _ = Products.objects.filter(
            manufacturer_identifier__in=manufacturer_ids
        ).delete()
        return (
            f"Deleted {deleted} products where manufacturer_id matches ids from file."
        )

    return "No matching products found or file format is incorrect."
